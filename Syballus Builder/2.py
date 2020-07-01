import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

excel = openpyxl.load_workbook("Çıktı.xlsx")
Sheet = excel.worksheets[2]
Hocalar = []
gSaat = 1
tGün = 1
Kök = 2

Uyarılar = []

Sheet = excel.worksheets[0]
while str(Sheet.cell(1,gSaat).value) != "None":
    gSaat += 1
gSaat -= 2
while str(Sheet.cell(tGün,1).value) != "None":
    tGün += 1
tGün -= 2
Sheet = excel.worksheets[2]
while str(Sheet.cell(Kök, 1).value) != "None":
    new = str(Sheet.cell(Kök, 1).value)
    HocaAdı = ""
    Sınıflar = []
    DersAdı = ""
    for i in range(len(new)):
        if new[i] == ",":
            HocaAdı = new[0:i]
            DersAdı = new[i+1: len(new)]
    i = 0
    while str(Sheet.cell(Kök+i, 2).value) != "None":
        Sınıflar.append([str(Sheet.cell(Kök+i, 2).value), str(Sheet.cell(Kök, 3).value)])
        i += 1
    Hocalar.append([HocaAdı,DersAdı,Sınıflar])
    Kök += len(Sınıflar)+1



Tek = True

#------------------------------Veri Analizi--------------------------------------
def BaşkaVarmı(HocaAdı, DersAdı,SınıfAdı, y):
    ws = excel.worksheets[1]
    Sınıf = 1
    DersAdı = DersAdı.replace(" ", "")
    HocaAdı = HocaAdı.replace(" ", "")

    while str(ws.cell(Sınıf,1).value) != SınıfAdı:
        Sınıf += 1

    for i in range(2, gSaat + 2):
        new = str(ws.cell(Sınıf + y, i).value)
        new = new.replace(" ", "")
        for b in range(len(new)):
            if new[b] == ",":
                if new[b+1:len(new)] == HocaAdı:
                    return False
    return True

def UygunMu(HocaAdı, SınıfAdı, y, x):

    ws = excel.worksheets[0]
    kök = 1
    while str(ws.cell(kök, 1).value) != HocaAdı:
        kök += 1
    new = str(ws.cell(kök+y, x).value)
    new = new.replace(" ", "")

    if new != "None":

        return False
    ws = excel.worksheets[1]
    kök = 1
    while str(ws.cell(kök, 1).value) != SınıfAdı:
        kök += 1
    new = str(ws.cell(kök+y, x).value)
    new = new.replace(" ", "")
    if new != "None":

        return False
    return True

def HocayaUygunMu(HocaAdı, y , x):

    ws = excel.worksheets[0]
    kök = 1
    while str(ws.cell(kök, 1).value) != HocaAdı:
        kök += 1
    new = str(ws.cell(kök + y, x).value)
    new = new.replace(" ", "")
    if new != "None":
        return False

def Yazıcı(HocaAdı, DersAdı, SınıfAdı, y, x1, x2):

    i = 1

    ws = excel.worksheets[0]
    if x2 > gSaat+1 or x1 > gSaat+1:

        return False
    for i in range(x1, x2+1):
        if UygunMu(HocaAdı, SınıfAdı, y, i) == False:

            return False
    kök = 1
    while str(ws.cell(kök,1).value) != HocaAdı:
        kök += 1
    for i in range(x1, x2+1):
        ws.cell(kök+y, i).value = SınıfAdı
    ws = excel.worksheets[1]
    kök = 1
    while str(ws.cell(kök, 1).value) != SınıfAdı:
        kök += 1
    for i in range(x1, x2+1):
        ws.cell(kök+y, i).value = "{}, {}".format(DersAdı, HocaAdı)
    return True

def DersDeğiştirici(HocaAdı, DersAdı, SınıfAdı, y, x):

    Sınıf = 1
    sy = 1
    sx = 2
    ws = excel.worksheets[1]
    Hedef = ""
    while str(ws.cell(Sınıf, 1).value) != SınıfAdı:
        Sınıf += 1
    sy = Sınıf + y
    sx = x
    Hedef = str(ws.cell(Sınıf+y, x).value)
    HedefSayısı = 0

    for i in range(2,gSaat+2):
        if str(ws.cell(Sınıf+y, i).value) == Hedef:
            HedefSayısı += 1

    for a in range(1,tGün+1):
        for b in range(2, gSaat+2):
            if a != y:
                if Yazıcı(HocaAdı, DersAdı, SınıfAdı, a, b, b+HedefSayısı-1):
                    for ç in range(sx, sx+HedefSayısı):
                        ws.cell(sy, ç).value = "None"
                    return True
    return False

def SorunÇözücü(HocaAdı, DersAdı, SınıfAdı, saat):

    Sınıf = 1
    mk = ""
    ws = excel.worksheets[1]
    while str(ws.cell(Sınıf, 1).value) != SınıfAdı:
        Sınıf += 1
    y = 1
    x = 2
    while y < tGün:
        while x < gSaat+1:
            Ps = True
            for i in range(x, x+saat+1):
                if HocayaUygunMu(HocaAdı, y, i) == False:
                    Ps = False
            Değişecekler = []
            for a in range(x, x+saat):
                if str(ws.cell(Sınıf+y,a).value) != "None":
                    new = str(ws.cell(Sınıf + y, a).value)
                else:
                    new = ""
                old = ""
                for char in range(len(new)):
                    if new[char] == ",":
                        old = new[0: char]
                        mk = old
                        new = new[char+2:len(new)]
                        break
                pss = True
                for c in Değişecekler:
                    if c == new:
                        pss = False
                if pss and len(new) > 0:
                    Değişecekler.append(new)
            if Ps == True and len(Değişecekler) == 1:
                if DersDeğiştirici(Değişecekler[0], mk, SınıfAdı, y, x):
                    return True
            x += 1
        y += 1
    return False



#------------------------------Fonksiyonlar--------------------------------------
dokuz = False
YeniHocalar = []


for Hoca in Hocalar:
    Dersler = []
    Sheet = excel.worksheets[1]
    for s in range(len(Hoca[2])):

        ilk = []
        İkinci = []
        new = Hoca[2][s][1]
        new = new.replace(" ", "")
        rt = 0
        for b in range(len(new)):
            if new[b] == "+":
                if Tek and int(new[rt:b]) % 2 == 1:
                    İkinci.append(new[rt:b])
                    rt = b+1
                elif Tek and int(new[rt:b]) % 2 != 1:
                    ilk.append(new[rt:b])
                    rt = b+1

                if Tek == False and int(new[rt:b]) % 2 == 1:
                    ilk.append(new[rt:b])
                    rt = b + 1
                elif Tek == False and int(new[rt:b]) % 2 != 1:
                    İkinci.append(new[rt:b])
                    rt = b + 1
            elif b == len(new) -1 :
                if Tek and int(new[rt:b+1]) % 2 == 1:
                    İkinci.append(new[rt:b+1])
                    rt = b+1
                elif Tek and int(new[rt:b+1]) % 2 != 1:
                    ilk.append(new[rt:b+1])
                    rt = b+1
                if Tek == False and int(new[rt:b+1]) % 2 != 1:
                    İkinci.append(new[rt:b+1])
                if Tek == False and int(new[rt:b+1]) % 2 == 1:
                    ilk.append(new[rt:b+1])

        x = 2
        y = 0
        if len(ilk) > 0:
            while y <= tGün:
                y += 1
                x = 2
                while x <= gSaat + 1:

                    if BaşkaVarmı(Hoca[0], Hoca[1], Hoca[2][s][0], y) and len(ilk)>0:
                        if Yazıcı(Hoca[0], Hoca[1], Hoca[2][s][0], y, x, x + int(ilk[0]) - 1):
                            x = 2
                            y = 1
                            ilk.pop(0)
                        else:
                            x += 1
                    else:
                        x += 1
        if len(ilk) > 0:
            for a in range(len(ilk)):
                if SorunÇözücü(Hoca[0], Hoca[1], Hoca[2][s][0], int(ilk[0])):
                    print("SorunÇözüldü")
                    ilk.pop(0)
        if len(ilk)>0:
            Uyarılar.append(ilk)


        YeniHocalar.append([Hoca[0], Hoca[1], Hoca[2][s][0], İkinci])

#--------------------------ilk------------------------------------------
print(YeniHocalar)
for Hoca in YeniHocalar:
    ders = Hoca[3]
    y = 1
    x = 2
    if len(ders) > 0:
        while y <= tGün:

            x = 2
            while x <= gSaat+1:
                if BaşkaVarmı(Hoca[0], Hoca[1], Hoca[2], y) and len(ders) > 0:
                    if Yazıcı(Hoca[0], Hoca[1], Hoca[2], y, x, x + int(ders[0]) - 1):
                        x = 2
                        y = 1
                        ders.pop(0)
                    else:
                        x += 1
                else:
                    x += 1
            y += 1
    if len(ders) > 0:
        for a in range(len(ders)):
            if SorunÇözücü(Hoca[0], Hoca[1], Hoca[2], int(ders[0])):

                ders.pop(0)
    if len(ders) > 0:
        Uyarılar.append(ders)





'''
dokuz = False
YeniHocalar = []

for Hoca in Hocalar:
minus = 0
    print("baş")
    Tur = 0
    if dokuz:
        break
    Hedef = ""

    TümSınıflar = len(Hoca[2])
    for i in range(0, TümSınıflar):
        print("Orta")
        print(Hoca)
        print(i)
        DersSayısı = []
        if dokuz:
            break
        new = Hoca[2][i][1]
        rt = 0
        for b in range(len(new)):
            minus = 0
            if dokuz:
                break
            if new[b] == "+":
                DersSayısı.append(new[rt: b])
                rt = b + 1
            elif b == len(new) - 1:
                DersSayısı.append(new[rt: b + 1])
        kök = 1
        Sheet = excel.worksheets[0]
        while str(Sheet.cell(kök, 1).value) != Hoca[0]:
            kök += 1
        minus = 0
        for ç in range(len(DersSayısı)):
            if Tek and int(DersSayısı[ç]) % 2 == 1:
                DersSayısı.append(DersSayısı[ç])
                DersSayısı.pop(ç)
                minus += 1
            elif Tek == False and int(DersSayısı[ç]) % 2 == 0:
                DersSayısı.append(DersSayısı[ç])
                DersSayısı.pop(ç)
                minus += 1

        while len(DersSayısı)-minus > 0:
            print(len(DersSayısı)-minus)
            if dokuz:
                break
            if Hoca[0] == "Rukiye" and Hoca[2][i][0] == "9/B":
                excel.save("Prog.xlsx")
                excel.close()
                dokuz = True
                break

            y = kök + 1
            x = 2
            if Tur == 0:
                try:
                    print(y,x)
                    while y < kök + tGün + 1 and len(DersSayısı)-minus > 0:
                        if BaşkaVarmı(Hoca[0], Hoca[1], Hoca[2][i][0], y):
                            if Yazıcı(Hoca[0], Hoca[1], Hoca[2][i][0], y - kök, x, x + int(DersSayısı[0]) - 1):

                                y = 1
                                x = 2
                            DersSayısı.pop(0)
                        elif x == gSaat + 1:
                            x = 2
                            y += 1
                        else:
                            x += 1
                    Tur += 1
                except:
                    pass
            elif Tur == 1 and len(DersSayısı)-minus > 0:
                if(SorunÇözücü(Hoca[0],Hoca[1],Hoca[2][i][0],int(DersSayısı[0]))):

                    Tur = 0
                    DersSayısı.pop(0)

'''
excel.save("Prog.xlsx")
excel.close()
