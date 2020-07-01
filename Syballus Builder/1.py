

#--------------------------------------------------------------------------------------------------







#Lütfen üsttekki kısımdan "Run" Bölümüne gelin ve oradan "Run Module" kısmını seçin.









#--------------------------------------------------------------------------------------------------


import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

excel = openpyxl.Workbook()
ActiveSheet = excel.active
Hocalar = []
Sınıf = []
Günler = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
thick_border = Border(left=Side(style='thick'),
                     right=Side(style='thick'),
                     top=Side(style='thick'),
                     bottom=Side(style='thick'))
thin_border = Border(left=Side(style='double'),
                     right=Side(style='double'),
                     top=Side(style='double'),
                     bottom=Side(style='double'))


read = openpyxl.load_workbook("1.xlsx")
readSheet = read.active
DersSaati = int(readSheet.cell(2, 4).value)
i = 2
while str(readSheet.cell(i,1).value) != "None":
    HocaAdı = str(readSheet.cell(i, 1).value)
    DersAdı = str(readSheet.cell(i, 2).value)
    Sınıflar = []
    b = 0
    new = str(readSheet.cell(i, 3).value)
    new = new.replace(" ", "")
    readSheet.cell(i, 3).value = new
    root = 0
    for b in range(len(str(readSheet.cell(i, 3).value))):

        if str(readSheet.cell(i, 3).value)[b] == ",":
            Sınıflar.append(str(readSheet.cell(i, 3).value)[root: b])
            root = b+1
        if b == len(str(readSheet.cell(i, 3).value))-1:
            Sınıflar.append(str(readSheet.cell(i, 3).value)[root: b+1])
    Hocalar.append([HocaAdı, DersAdı, Sınıflar])
    i += 1

i = 2

while str(readSheet.cell(i, 3).value) != "None":
    root = 0
    new = str(readSheet.cell(i, 3).value)
    new = new.replace(" ", "")
    for b in range(len(new)):
        if new[b] == ",":
            Pass = True
            for cl in Sınıf:
                if new[root: b] == cl:
                    Pass = False
            if Pass:
                Sınıf.append(new[root: b])
            root = b+1
        if b == len(new)-1:
            Pass = True
            for cl in Sınıf:
                if new[root: b+1] == cl:
                    Pass = False
            if Pass:
                Sınıf.append(new[root: b+1])
    i += 1


for i in range(5, 12):

    if str(readSheet.cell(2, i).value) != "None":
        Günler.remove(readSheet.cell(1, i).value)

#-----------------Veri İşleme--------------------------

kök = 1
ActiveSheet.column_dimensions["A"].width = 15
for i in range(0,len(Hocalar)):
    ActiveSheet.cell(kök, 1).value = Hocalar[i][0]
    ActiveSheet.cell(kök, 1).font = Font(bold=True,color="FF1717")
    ActiveSheet.cell(kök, 1).border = thick_border
    for b in range(1,DersSaati+1):
        ActiveSheet.cell(kök, 1+b).value = "{}. ders".format(b)
        ActiveSheet.cell(kök, 1 + b).border = thick_border
    for b in range(len(Günler)):
        ActiveSheet.cell(kök+1+b, 1).value = Günler[b]
        ActiveSheet.cell(kök + 1 + b, 1).border = thick_border
    for a in range(1,DersSaati+1):
        for b in range(1,len(Günler)+1):
            ActiveSheet.cell(kök+b, a+1).border = thick_border
    kök += len(Günler)+3

#--------------Hoca Programları--------------------

ActiveSheet = excel.create_sheet("Sınıf Programları", 1)
ActiveSheet.column_dimensions["A"].width = 15
kök = 1
for i in range(0,len(Sınıf)):
    ActiveSheet.cell(kök, 1).value = Sınıf[i]
    ActiveSheet.cell(kök, 1).font = Font(bold=True, color="FF1717")
    ActiveSheet.cell(kök, 1).border = thick_border
    for b in range(1,DersSaati+1):
        ActiveSheet.cell(kök, 1+b).value = "{}. ders".format(b)
        ActiveSheet.cell(kök, 1 + b).border = thick_border
    for b in range(len(Günler)):
        ActiveSheet.cell(kök+1+b, 1).value = Günler[b]
        ActiveSheet.cell(kök + 1 + b, 1).border = thick_border
    for a in range(1,DersSaati+1):
        for b in range(1,len(Günler)+1):
            ActiveSheet.cell(kök+b,a+1).border = thick_border
    kök += len(Günler)+3


#-------------------------------------3. Sheet--------------------------------
ActiveSheet = excel.create_sheet("Hocalar ve Sınıfları", 2)
ActiveSheet.cell(1, 1).value = "Hoca Adı, Dersinin adı"
#ActiveSheet.cell(1, 1).border = thin_border
ActiveSheet.cell(1, 2).value = "Sınıf Adı"
#ActiveSheet.cell(1, 2).border = thin_border
ActiveSheet.cell(1, 3).value = "Sınıfa verilecek ders saati"
#ActiveSheet.cell(1, 3).border = thin_border
kök = 2
ActiveSheet.column_dimensions["C"].width = 25
ActiveSheet.column_dimensions["A"].width = 25
for i in range(len(Hocalar)):
    ActiveSheet.cell(kök, 1).value = "{}, {}".format(Hocalar[i][0], Hocalar[i][1])
    ActiveSheet.cell(kök, 1).font = Font(bold=True, color="FF1717")
    ActiveSheet.cell(kök, 1).border = thin_border
    for b in range(len(Hocalar[i][2])):
        ActiveSheet.cell(kök+b, 2).value = Hocalar[i][2][b]
        ActiveSheet.cell(kök+b, 2).border = thin_border
    kök += len(Hocalar[i][2])+1


excel.save("Çıktı.xlsx")
print("--------------------------\nİşleminiz Başarıyla Gerçekleşti\n--------------------------")

excel.close()
